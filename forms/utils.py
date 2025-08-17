import logging
from django.http import HttpResponse
import openpyxl
from requests import Response

import csv
from rest_framework import status
from .mailersend import send_mailersend_email
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from django.apps import apps
import os

# Configuración del log
logger = logging.getLogger(__name__)

info_questions = {
    "userName": "Nombre",
    "selectForm": "Tipo de Análisis",
    "identificationType": "Tipo de Documento", # Nesesario para la busqueda y results
    "identificationNumber": "Número de Identificación",
    "email": "Correo Electrónico",
    "marketReach": "Tamaño del de la Empresa",
}



# Función para enviar correos de diagnóstico con mailersend
def send_diagnostic_email(to_email, subject, context):
        from .models import CompletedForm
        
        catalog_url = "https://rcklean.wirkconsulting.com/catalogo"
        portfolio_url = "https:/acsalazar.com"

        
            # === Email to client ===
        results_url = f"https://rcklean.wirkconsulting.com/results/" # Pendiente de Heroku Student plan 
        subject_client = "Epa! has realizado un test en nustra herramienta de autodiagnóstico"
        text_client = f"""
        Hello {CompletedForm.user}


        Gracias por completar el test de autodiagnóstico. Aquí tienes un resumen de tus resultados:
        

        Completaste el Test : {CompletedForm.form_title}


        Puedes ver los detalles de tu diagnóstico en el siguiente enlace: {results_url}
        """

        html_client = f"""
        <html><body style="font-family: Arial;">
          <h2>Vuelve Pornto, {CompletedForm.user}!</h2>
          <p>Actualziamos nuestro Catalogo de Evaluaciones Constattemente, vuelve para encotrar mas contenido:</p>
          <p><strong>Comletaste el test:</strong> {CompletedForm.form_title}</p>
          <p><strong>Nuestro catalogo: </strong> ${catalog_url}</p>
          <p><strong>Eres siemrpre bienvenid@:</strong></p>
          <a href="{results_url}" style="display:inline-block;padding:10px 20px;background-color:#28a745;color:#fff;text-decoration:none;border-radius:4px;">
            Puedes consultar los resultados de este y otras evaluaicones de nuestra plataforma aquí:
          </a>
          <p style="margin-top:20px;font-size:12px;color:#888;">by acsalazar {portfolio_url} · acsalazar-19@hotmail.com · Medellín, Colombia</p>
        </body></html>
        """

        send_mailersend_email(CompletedForm.email, subject_client, text_client, html_client)

        # === Email to business ===
        subject_admin = "🚨 Un usuario ha completado un test"
        text_admin = f"""
        New client request:

        Name: {CompletedForm.user}
        Email: {CompletedForm.email}
        Form Title: {CompletedForm.form_title}



        """

        html_admin = f"""
        <html><body style="font-family: Arial;">
          <h2>🚨 New Quote Request Received</h2>
          <p><strong>Name:</strong> {CompletedForm.user}</p>
          <p><strong>Email:</strong> {CompletedForm.email}</p>
          <p><strong>Form Title:</strong> {CompletedForm.form_title}</p>
          <a href="{results_url}" style="display:inline-block;padding:10px 20px;background-color:#007bff;color:#fff;text-decoration:none;border-radius:4px;">View Estimate</a>
        </body></html>
        """

        send_mailersend_email("acsalazar-19@hotmail.com", subject_admin, text_admin, html_admin)

        return Response({ "Texto de confirmación enviado": "El correo de confirmación ha sido enviado correctamente."
        }, status=status.HTTP_200_OK)



# Función para encontrar el plan de trabajo correspondiente a una categoría específica
def find_plan(category_id, average):
    Category = apps.get_model('forms', 'Category')
    category = Category.objects.get(id=category_id)
    
    # Obtener los niveles y planes para la categoría
    levels = category.levels.all()

    # Determinar el nivel adecuado basado en el puntaje promedio
    if average < 2:
        level_needed = 1
    elif average < 3:
        level_needed = 2
    elif average < 4:
        level_needed = 3
    else:
        level_needed = 4

    # Buscar el plan adecuado en los niveles de la categoría
    for level in levels:
        if level.level == level_needed:
            plan = level.plans.first()  # Obtener el primer plan disponible para este nivel
            return plan.text if plan else "No hay plan disponible para este nivel"
    
    return "No hay plan disponible para esta categoría"

# Función para calcular los promedios de categorías a partir de las respuestas
def calculate_category_averages(answers):
    try:
        category_scores = {}
        category_counts = {}
        categories_details = {}

        for answer in answers:
            category = answer['category']
            value = answer['value']
            
            if category['name'] == 'Complejidad':
                continue

            category_name = category['name']

            if category_name in category_scores:
                category_scores[category_name] += value
                category_counts[category_name] += 1
            else:
                category_scores[category_name] = value
                category_counts[category_name] = 1
                categories_details[category_name] = category

        category_averages = []
        for category_name in category_scores:
            average = category_scores[category_name] / category_counts[category_name]
            category = categories_details[category_name]
            category_id = category['id']
            plan_to_show = find_plan(category_id, average)

            category_averages.append({
                'category': category,
                'average': average,
                'plan': plan_to_show
            })

# El prompt general usara palabras clave como bajo, medio y alto para determinar el nivel conocimuiento del usuario y segun ello
# Recomienda documentacion y videos, o cursos gratuitos curso personalisado ( link a marketplace).
        def prompt_context_level(prompt, level_base):

            level_base = ['Bajo', 'Medio', 'Alto']
            prompt = "Recomienda videos y articulos y brinda informacion en forma de plan de mejora basado en la documentación que tienes del tema del cuestionario ."
            if category_averages.average < 2:
                return f"{prompt} {level_base[0]} "
            elif category_averages.average < 3:
                return f"{prompt} {level_base[1]} "
            elif category_averages.average < 4:
                return f"{prompt} {level_base[2]} "
            
            return prompt_context_level(prompt, level_base)



        return category_averages
    except Exception as e:
        logging.error(f"Error calculating category averages: {e}")
        return []
    


    
# Función para obtener las preguntas ordenadas
def get_ordered_answer_questions(forms):
    ordered_answer_questions = []
    seen_questions = set()
    
    for form in forms:
        answers = form.content.get('answers', [])
        for ans in answers:
            question_text = ans['questionText']
            if question_text not in seen_questions:
                ordered_answer_questions.append(question_text)
                seen_questions.add(question_text)
                
    return ordered_answer_questions

# Función para construir una fila de datos
def build_row(obj, ordered_answer_questions, category_averages):
    row = [obj.user, obj.form_title, obj.created_at.strftime("%Y-%m-%d %H:%M:%S")]
    info_data = obj.content.get('info', {})
    for key in info_questions:
        row.append(info_data.get(key, "N/A"))
    answers_data = {ans['questionText']: ans.get('value', 'N/A') for ans in obj.content.get('answers', [])}
    for question in ordered_answer_questions:
        row.append(answers_data.get(question, "N/A"))
    
    for avg in category_averages:
        row.append(avg['plan'])
    return row

# Función para exportar datos como CSV
def export_as_csv(queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=completed_forms.csv'
    writer = csv.writer(response)

    if not queryset.exists():
        writer.writerow(['No hay datos para mostrar.'])
        return response

    forms_by_title = {}
    for obj in queryset:
        forms_by_title.setdefault(obj.form_title, []).append(obj)

    for form_title, forms in forms_by_title.items():
        writer.writerow([form_title])
        ordered_answer_questions = get_ordered_answer_questions(forms)
        headers = ['User', 'Form Title', 'Created At'] + list(info_questions.values()) + ordered_answer_questions
        
        all_category_averages = set()
        for form in forms:
            category_averages = calculate_category_averages(form.content.get('answers', []))
            all_category_averages.update(f"Autodiagnóstico {avg['category']['name']}" for avg in category_averages)
        headers += list(all_category_averages)

        writer.writerow(headers)

        for obj in forms:
            category_averages = calculate_category_averages(obj.content.get('answers', []))
            row = build_row(obj, ordered_answer_questions, category_averages)
            writer.writerow(row)

    return response

# Función para exportar datos como Excel
def export_as_excel(queryset):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=completed_forms.xlsx'
    wb = openpyxl.Workbook()

    if not queryset.exists():
        ws = wb.active
        ws.append(['No hay datos para mostrar.'])
        wb.save(response)
        return response

    forms_by_title = {}
    for obj in queryset:
        forms_by_title.setdefault(obj.form_title, []).append(obj)

    for form_title, forms in forms_by_title.items():
        ws = wb.create_sheet(title=form_title[:31])
        ordered_answer_questions = get_ordered_answer_questions(forms)
        headers = ['User', 'Form Title', 'Created At'] + list(info_questions.values()) + ordered_answer_questions
        
        all_category_averages = set()
        for form in forms:
            category_averages = calculate_category_averages(form.content.get('answers', []))
            all_category_averages.update(f"Autodiagnóstico {avg['category']['name']}" for avg in category_averages)
        headers += list(all_category_averages)

        ws.append(headers)

        for obj in forms:
            category_averages = calculate_category_averages(obj.content.get('answers', []))
            row = build_row(obj, ordered_answer_questions, category_averages)
            ws.append(row)

        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save(response)
    return response


 